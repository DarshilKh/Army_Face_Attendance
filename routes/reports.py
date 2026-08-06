from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from models.database import db, Employee, Attendance, User
from datetime import datetime, timedelta, date as dt_date
from utils.logger import app_logger
import io
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def index():
    """Reports dashboard"""
    try:
        today = dt_date.today()

        today_attendance = Attendance.query.filter(
            Attendance.date == today
        ).count()

        total_employees = Employee.query.filter_by(is_active=True).count()

        month_start = today.replace(day=1)
        month_attendance = Attendance.query.filter(
            Attendance.date >= month_start
        ).count()

        stats = {
            'today_present': today_attendance,
            'total_employees': total_employees,
            'today_absent': total_employees - today_attendance,
            'month_total': month_attendance
        }

        units = db.session.query(Employee.unit).distinct().filter(
            Employee.unit.isnot(None), Employee.unit != ''
        ).all()
        units = sorted([u[0] for u in units if u[0]])

        return render_template('reports.html', stats=stats, units=units)
    except Exception as e:
        app_logger.error(f"Error loading reports page: {e}", exc_info=True)
        return render_template('error.html', error_message=str(e)), 500


@reports_bp.route('/analytics')
@login_required
def analytics():
    """Advanced analytics page (placeholder)"""
    units = db.session.query(Employee.unit).distinct().filter(
        Employee.unit.isnot(None), Employee.unit != ''
    ).all()
    units = sorted([u[0] for u in units if u[0]])
    return render_template('reports.html', units=units)


@reports_bp.route('/audit-logs')
@login_required
def audit_logs():
    """Audit logs page"""
    from models.database import AuditLog
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(100).all()
    return render_template('reports.html', audit_logs=logs)


# ── Fix 1: generate_report() now branches on format instead of always
#    calling download_daily_excel() regardless of what the user selected.
# ── Fix 2: download_daily_pdf() is now implemented (was imported but
#    never written — reportlab was a dead import before this change).
# ─────────────────────────────────────────────────────────────────────

@reports_bp.route('/generate', methods=['POST'])
@login_required
def generate_report():
    """
    Generate and download attendance report.

    Reads 'format' from the JSON body ('excel' or 'pdf') and routes to
    the correct download helper, passing along the date range, unit
    filter, and report_type selected on the form.
    """
    try:
        data           = request.get_json()
        report_type    = data.get('format', 'excel')   # 'excel' or 'pdf'
        start_date_str = data.get('start_date')
        end_date_str   = data.get('end_date') or start_date_str
        unit           = (data.get('unit') or '').strip() or None
        range_label    = data.get('report_type', 'custom')  # daily/weekly/monthly/custom

        # ── Branching fix ─────────────────────────────────────────────────
        if report_type == 'pdf':
            return download_daily_pdf(start_date_str, end_date_str, unit, range_label)
        else:
            return download_daily_excel(start_date_str, end_date_str, unit, range_label)
        # ─────────────────────────────────────────────────────────────────

    except Exception as e:
        app_logger.error(f"Error generating report: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/daily', methods=['GET'])
@login_required
def daily_report():
    """Generate daily attendance report"""
    try:
        date_str    = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()

        attendance_records = db.session.query(Attendance, Employee).join(
            Employee, Attendance.employee_id == Employee.id
        ).filter(Attendance.date == report_date).all()

        all_employees = Employee.query.filter_by(is_active=True).all()

        report_data = []
        for emp in all_employees:
            attendance = next(
                (a for a, e in attendance_records if a.employee_id == emp.id), None
            )
            report_data.append({
                'army_id':      emp.army_id,
                'name':         emp.full_name,
                'rank':         emp.rank or 'N/A',
                'unit':         emp.unit or 'N/A',
                'check_in':     attendance.check_in_time.strftime('%I:%M %p')
                                if attendance and attendance.check_in_time
                                else 'Absent',
                'check_out':    attendance.check_out_time.strftime('%I:%M %p')
                                if attendance and attendance.check_out_time
                                else '-',
                'status':       'Present' if attendance else 'Absent',
                'status_class': 'success' if attendance else 'danger'
            })

        total_employees       = len(all_employees)
        present_count         = len([r for r in report_data if r['status'] == 'Present'])
        absent_count          = total_employees - present_count
        attendance_percentage = (present_count / total_employees * 100) if total_employees > 0 else 0

        stats = {
            'total':      total_employees,
            'present':    present_count,
            'absent':     absent_count,
            'percentage': round(attendance_percentage, 1)
        }

        return render_template('reports.html',
                               report_data=report_data,
                               report_date=report_date,
                               stats=stats)

    except Exception as e:
        app_logger.error(f"Error generating daily report: {e}", exc_info=True)
        return render_template('error.html', error_message=str(e)), 500


@reports_bp.route('/monthly')
@login_required
def monthly_report():
    """Generate monthly attendance report"""
    try:
        month            = request.args.get('month', datetime.now().strftime('%Y-%m'))
        year, month_num  = map(int, month.split('-'))

        start_date = datetime(year, month_num, 1)
        end_date   = datetime(year + 1, 1, 1) if month_num == 12 else datetime(year, month_num + 1, 1)

        attendance_records = Attendance.query.filter(
            Attendance.date >= start_date.date(),
            Attendance.date <  end_date.date()
        ).all()

        employees    = Employee.query.filter_by(is_active=True).all()
        working_days = (end_date - start_date).days
        report_data  = []

        for emp in employees:
            emp_attendance = [a for a in attendance_records if a.employee_id == emp.id]
            present_days   = len(set(a.date for a in emp_attendance))
            percentage     = round(present_days / working_days * 100, 1) if working_days > 0 else 0

            report_data.append({
                'army_id':      emp.army_id,
                'name':         emp.full_name,
                'rank':         emp.rank or 'N/A',
                'unit':         emp.unit or 'N/A',
                'present_days': present_days,
                'working_days': working_days,
                'percentage':   percentage,
                'status_class': 'success' if percentage >= 90 else 'warning' if percentage >= 75 else 'danger'
            })

        report_data.sort(key=lambda x: x['percentage'], reverse=True)

        return render_template('reports.html',
                               report_data=report_data,
                               month=month,
                               month_name=start_date.strftime('%B %Y'))

    except Exception as e:
        app_logger.error(f"Error generating monthly report: {e}", exc_info=True)
        return render_template('error.html', error_message=str(e)), 500


@reports_bp.route('/employee/<army_id>')
@login_required
def employee_report(army_id):
    """Individual employee report"""
    try:
        employee = Employee.query.filter_by(army_id=army_id).first_or_404()

        start_date_str = request.args.get(
            'start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        )
        end_date_str = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))

        start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end   = datetime.strptime(end_date_str,   '%Y-%m-%d').date()

        attendance_records = Attendance.query.filter(
            Attendance.employee_id == employee.id,
            Attendance.date >= start,
            Attendance.date <= end
        ).order_by(Attendance.date.desc()).all()

        present_days = len(set(a.date for a in attendance_records))
        total_days   = (end - start).days + 1
        percentage   = round(present_days / total_days * 100, 1) if total_days > 0 else 0

        stats = {
            'present_days': present_days,
            'total_days':   total_days,
            'percentage':   percentage,
            'on_time':      sum(1 for a in attendance_records if a.status == 'present'),
            'late':         sum(1 for a in attendance_records if a.status == 'late'),
            'half_day':     sum(1 for a in attendance_records if a.status == 'half_day')
        }

        return render_template('reports.html',
                               employee=employee,
                               attendance_records=attendance_records,
                               stats=stats,
                               start_date=start,
                               end_date=end)

    except Exception as e:
        app_logger.error(f"Error generating employee report: {e}", exc_info=True)
        return render_template('error.html', error_message=str(e)), 500


@reports_bp.route('/download/excel/daily/<date_str>')
@login_required
def _fetch_report_records(start_date, end_date, unit=None):
    """
    Shared query helper: attendance records for active employees in
    [start_date, end_date], optionally filtered to one unit.
    Returns (records, employee_count) where records is a list of
    (Attendance, Employee) tuples ordered by date then name.
    """
    query = db.session.query(Attendance, Employee).join(
        Employee, Attendance.employee_id == Employee.id
    ).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date,
        Employee.is_active == True
    )
    if unit:
        query = query.filter(Employee.unit == unit)

    records = query.order_by(Attendance.date.asc(), Employee.full_name.asc()).all()

    employee_count_query = Employee.query.filter_by(is_active=True)
    if unit:
        employee_count_query = employee_count_query.filter_by(unit=unit)
    employee_count = employee_count_query.count()

    return records, employee_count


def download_daily_excel(start_date_str, end_date_str=None, unit=None, range_label='custom'):
    """
    Download attendance report as Excel for a date range (a single-day
    range when start_date == end_date), optionally filtered to a unit.
    """
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else start_date
        is_single_day = start_date == end_date

        records, employee_count = _fetch_report_records(start_date, end_date, unit)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        last_col = 7
        ws.merge_cells(f'A1:{openpyxl.utils.get_column_letter(last_col)}1')
        title_cell = ws['A1']
        if is_single_day:
            title_cell.value = f'{range_label.title()} Attendance Report - {start_date.strftime("%d %B %Y")}'
        else:
            title_cell.value = (
                f'{range_label.title()} Attendance Report - '
                f'{start_date.strftime("%d %b %Y")} to {end_date.strftime("%d %b %Y")}'
            )
        title_cell.font      = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        if unit:
            ws.merge_cells(f'A2:{openpyxl.utils.get_column_letter(last_col)}2')
            unit_cell = ws['A2']
            unit_cell.value = f'Unit: {unit}'
            unit_cell.alignment = Alignment(horizontal='center')

        header_fill      = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font      = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ['Date', 'Army ID', 'Name', 'Rank', 'Unit', 'Check In', 'Check Out']
        for col, header in enumerate(headers, 1):
            cell           = ws.cell(row=4, column=col, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_alignment

        row = 5
        for att, emp in records:
            ws.cell(row=row, column=1, value=att.date.strftime('%d-%b-%Y'))
            ws.cell(row=row, column=2, value=emp.army_id)
            ws.cell(row=row, column=3, value=emp.full_name)
            ws.cell(row=row, column=4, value=emp.rank or 'N/A')
            ws.cell(row=row, column=5, value=emp.unit or 'N/A')
            ws.cell(row=row, column=6,
                    value=att.check_in_time.strftime('%I:%M %p') if att.check_in_time else '-')
            ws.cell(row=row, column=7,
                    value=att.check_out_time.strftime('%I:%M %p') if att.check_out_time else '-')
            row += 1

        # Summary block below the data
        working_days = (end_date - start_date).days + 1
        row += 1
        ws.cell(row=row, column=1, value='Employees:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=employee_count)
        row += 1
        ws.cell(row=row, column=1, value='Days in range:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=working_days)
        row += 1
        ws.cell(row=row, column=1, value='Attendance records:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(records))

        for col in range(1, last_col + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename_suffix = start_date_str if is_single_day else f'{start_date_str}_to_{end_date_str}'
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{range_label}_report_{filename_suffix}.xlsx'
        )

    except Exception as e:
        app_logger.error(f"Error generating Excel: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


# ── Fix 2: PDF generation — reportlab was fully imported at the top of the
#    file in the original code but this function was never written, making
#    the import dead code and PDF downloads impossible.
# ─────────────────────────────────────────────────────────────────────────────
@reports_bp.route('/download/pdf/daily/<date_str>')
@login_required
def download_daily_pdf(date_str, end_date_str=None, unit=None, range_label='custom'):
    """
    Download attendance report as PDF for a date range (a single-day
    range when start_date == end_date), optionally filtered to a unit.

    When hit directly via the GET route, end_date/unit/report_type can
    be supplied as query params; generate_report() calls this function
    directly with explicit arguments for the range selected on the form.
    """
    try:
        if end_date_str is None:
            end_date_str = request.args.get('end_date', date_str)
        if unit is None:
            unit = request.args.get('unit') or None
        range_label = request.args.get('report_type', range_label)

        start_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        is_single_day = start_date == end_date

        records, employee_count = _fetch_report_records(start_date, end_date, unit)

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            fontSize=16
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            alignment=TA_CENTER,
            fontSize=10
        )

        elements = []

        # Title
        if is_single_day:
            title_text = f'{range_label.title()} Attendance Report - {start_date.strftime("%d %B %Y")}'
        else:
            title_text = (
                f'{range_label.title()} Attendance Report - '
                f'{start_date.strftime("%d %b %Y")} to {end_date.strftime("%d %b %Y")}'
            )
        elements.append(Paragraph(title_text, title_style))
        if unit:
            elements.append(Paragraph(f'Unit: {unit}', subtitle_style))
        elements.append(Spacer(1, 20))

        # Table data
        table_data = [['Date', 'Army ID', 'Name', 'Rank', 'Unit', 'Check In', 'Check Out']]

        for att, emp in records:
            table_data.append([
                att.date.strftime('%d-%b-%Y'),
                emp.army_id,
                emp.full_name,
                emp.rank or 'N/A',
                emp.unit or 'N/A',
                att.check_in_time.strftime('%I:%M %p') if att.check_in_time else '-',
                att.check_out_time.strftime('%I:%M %p') if att.check_out_time else '-',
            ])

        # Table styling
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor('#1F4788')),
            ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',       (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
            ('GRID',           (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ]))

        elements.append(table)

        working_days = (end_date - start_date).days + 1
        elements.append(Spacer(1, 16))
        elements.append(Paragraph(
            f'Employees: {employee_count} &nbsp;&nbsp; '
            f'Days in range: {working_days} &nbsp;&nbsp; '
            f'Attendance records: {len(records)}',
            subtitle_style
        ))

        doc.build(elements)
        buffer.seek(0)

        filename_suffix = date_str if is_single_day else f'{date_str}_to_{end_date_str}'
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'{range_label}_report_{filename_suffix}.pdf'
        )

    except Exception as e:
        app_logger.error(f"Error generating PDF: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500